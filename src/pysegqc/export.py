"""
Excel export and formatting functions for radiomics clustering results.

This module handles exporting clustering results to Excel with conditional
formatting, creating per-cluster training sheets, and training selection sheets.
"""

import logging
from typing import List, Optional, Dict, Any
from pathlib import Path
import numpy as np
import pandas as pd
import joblib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from .utils import CLUSTER_COLORS
from .visualization import plot_training_case_selection

logger = logging.getLogger(__name__)


def export_results(metadata_df, features_df, pca_data, labels, pca_model,
                   feature_names, output_dir, excel_path, scaler=None, impute_strategy='median',
                   training_selection_data=None, qa_results=None):
    """
    Export clustering results to Excel and CSV files with enhanced visualization.

    Args:
        training_selection_data: Optional dict from select_multi_structure_training_cases()
                                 If provided, adds Training_Case_Selection sheet
    """
    print(f"\n{'='*70}")
    print("Exporting Results")
    print(f"{'='*70}")

    # Create results dataframe
    results_df = metadata_df.copy()
    results_df['Cluster'] = labels

    # Add PC scores
    n_components = pca_data.shape[1]
    for i in range(min(10, n_components)):
        results_df[f'PC{i+1}'] = pca_data[:, i]

    # Add original features
    results_df = pd.concat([results_df, features_df], axis=1)

    # Sort by cluster number for better visualization
    results_df = results_df.sort_values('Cluster').reset_index(drop=True)

    # Prepare Excel file with multiple sheets
    output_excel = output_dir / f"{Path(excel_path).stem}_clustered.xlsx"
    n_clusters = labels.max() + 1

    # Create Excel writer
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # 1. Create Summary Sheet
        print("  Creating summary sheet...")
        summary_df = generate_cluster_summary(
            results_df,
            feature_names=feature_names,
            cluster_column='Cluster',
            n_top_features=10,
            include_confidence=False
        )
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # 2. Write main clustered data sheet
        print("  Writing clustered data sheet...")
        results_df.to_excel(writer, sheet_name='Clustered_Data', index=False)

        # 3. Write separate sheets per cluster
        print("  Creating per-cluster sheets...")
        for cluster in range(n_clusters):
            cluster_data = results_df[results_df['Cluster'] == cluster].copy()
            sheet_name = f'Cluster_{cluster}'
            cluster_data.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"    ✓ Sheet '{sheet_name}': {len(cluster_data)} samples")

        # 4. Write PCA Loadings sheet (replaces pca_loadings.csv)
        print("  Creating PCA Loadings sheet...")
        loadings_df = pd.DataFrame(
            pca_model.components_.T,
            columns=[f'PC{i+1}' for i in range(n_components)],
            index=feature_names
        )
        loadings_df.insert(0, 'Feature', feature_names)
        loadings_df.to_excel(writer, sheet_name='PCA_Loadings', index=False)

        # 5. Write Cluster Statistics sheet (replaces cluster_statistics.csv)
        print("  Creating Cluster Statistics sheet...")
        cluster_stats = []
        for cluster in range(n_clusters):
            mask = labels == cluster
            cluster_features = features_df.loc[results_df.index[results_df['Cluster'] == cluster]]

            stats = {
                'Cluster': cluster,
                'N_Samples': int(mask.sum()),
                'Pct_Total': f"{mask.sum() / len(labels) * 100:.1f}%"
            }

            if len(cluster_features) > 1:
                feature_means = cluster_features.mean()
                feature_stds = cluster_features.std()
                feature_vars = cluster_features.var()
                top_features = feature_vars.nlargest(min(10, len(feature_vars))).index

                for feat in top_features:
                    stats[f'{feat}_mean'] = feature_means[feat]
                    stats[f'{feat}_std'] = feature_stds[feat]

            cluster_stats.append(stats)

        cluster_stats_df = pd.DataFrame(cluster_stats)
        cluster_stats_df.to_excel(writer, sheet_name='Cluster_Statistics', index=False)

        # 6. Write Training_Case_Selection sheet if data provided
        if training_selection_data is not None:
            print("  Creating Training_Case_Selection sheet...")
            _create_training_selection_sheet(writer, training_selection_data)

    # Create PCA plot for selected training cases (after Excel is closed)
    if training_selection_data is not None:
        # Handle both new per-cluster format and legacy format
        if 'per_cluster_selections' in training_selection_data:
            # New format: Create a plot for all selected cases across all clusters
            all_selected_case_ids = []
            for cluster_info in training_selection_data['per_cluster_selections'].values():
                all_selected_case_ids.extend(cluster_info['selected_case_ids'])
            plot_training_case_selection(all_selected_case_ids, pca_data, labels, metadata_df, output_dir)
        else:
            # Legacy format
            selected_case_ids = training_selection_data['selected_case_ids']
            plot_training_case_selection(selected_case_ids, pca_data, labels, metadata_df, output_dir)

    # Apply conditional formatting
    print("  Applying conditional formatting...")
    _apply_conditional_formatting(output_excel, n_clusters)

    print(f"✓ Saved enhanced clustered data: {output_excel}")
    print(f"  - Summary sheet with cluster statistics")
    print(f"  - PCA_Loadings and Cluster_Statistics sheets")
    print(f"  - Main data sorted by cluster with color coding")
    print(f"  - {n_clusters} separate sheets (one per cluster)")

    # Save trained models for future predictions
    if scaler is not None:
        print(f"\n  Saving trained models for future predictions...")

        # Calculate cluster centroids in PCA space
        centroids = np.zeros((n_clusters, pca_data.shape[1]))
        for cluster in range(n_clusters):
            mask = labels == cluster
            centroids[cluster] = pca_data[mask].mean(axis=0)

        # Bundle trained components
        trained_models = {
            'scaler': scaler,
            'pca': pca_model,
            'centroids': centroids,
            'feature_names': feature_names,
            'n_clusters': n_clusters,
            'impute_strategy': impute_strategy,
            'metadata_columns': metadata_df.columns.tolist(),
            'n_components': pca_data.shape[1],
            # Training data for prediction visualization
            'training_pca_data': pca_data,
            'training_labels': labels,
            'training_metadata': metadata_df.copy(),
        }

        # Save QA thresholds so prediction mode can reproduce verdicts
        if qa_results is not None:
            trained_models['qa_thresholds'] = {
                'distance_sigma': 2.0,
                'iforest_contamination': 0.1,
                'per_cluster_stats': qa_results.get('per_cluster_stats', {}),
            }

        # Save to pickle file
        model_path = output_dir / 'trained_models.pkl'
        joblib.dump(trained_models, model_path)
        print(f"  ✓ Saved trained models: {model_path}")
        print(f"    - Scaler, PCA model, and {n_clusters} cluster centroids")
        print(f"    - Use with --predict flag to classify new cases")


def _create_per_cluster_training_sheets(writer, training_selection_data):
    """
    Create separate Training_Cluster_N sheets for each cluster's training selection.

    Args:
        writer: pd.ExcelWriter object
        training_selection_data: dict with per_cluster_selections
    """
    per_cluster_selections = training_selection_data['per_cluster_selections']
    n_clusters = len(per_cluster_selections)

    print(f"  Creating per-cluster training selection sheets...")

    # Create summary sheet first
    summary_data = []
    for cluster_id, cluster_info in sorted(per_cluster_selections.items()):
        summary_data.append({
            'Cluster_ID': cluster_id,
            'Cluster_Size': cluster_info['cluster_size'],
            'Cases_Selected': cluster_info['n_selected'],
            'Selection_Strategy': cluster_info['selection_strategy']
        })

    summary_df = pd.DataFrame(summary_data)

    # Write summary to Training_Summary sheet
    summary_df.to_excel(writer, sheet_name='Training_Summary', index=False)
    print(f"    ✓ Training_Summary: Overview of all {n_clusters} clusters")

    # Create separate sheet for each cluster's training selection
    for cluster_id, cluster_info in sorted(per_cluster_selections.items()):
        sheet_name = f'Training_Cluster_{cluster_id}'

        # Build dataframe from case_details
        cases_data = []
        for case_id in cluster_info['selected_case_ids']:
            case_metadata = cluster_info['case_details'][case_id]
            cases_data.append(case_metadata)

        cases_df = pd.DataFrame(cases_data)

        # Sort by distance from centroid (closest first)
        if 'Distance_From_Centroid' in cases_df.columns:
            cases_df = cases_df.sort_values('Distance_From_Centroid')
            cases_df['Distance_From_Centroid'] = cases_df['Distance_From_Centroid'].round(4)

        # Write to Excel
        cases_df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"    ✓ {sheet_name}: {len(cases_df)} training cases")


def _extract_case_metadata(case_id, metadata_df):
    """Extract patient identifiers from metadata for a given case."""
    result = {}
    if metadata_df is not None and case_id in metadata_df.index:
        for col in ('MRN', 'Plan_ID', 'Session_ID'):
            if col in metadata_df.columns:
                result[col] = metadata_df.loc[case_id, col]
    return result


def _create_training_selection_sheet(writer, training_selection_data):
    """
    Create Training_Case_Selection sheets with selection details.

    Supports three formats:
    - New per-cluster format: Creates separate Training_Cluster_N sheet for each cluster
    - Legacy concat mode: Single sheet with cluster_info
    - Legacy multi-structure mode: Single sheet with structure_details

    Args:
        writer: pd.ExcelWriter object
        training_selection_data: dict from select_training_cases_from_clustering() or select_multi_structure_training_cases()
    """
    # Detect format type
    if 'per_cluster_selections' in training_selection_data:
        # NEW FORMAT: Per-cluster separate sheets
        _create_per_cluster_training_sheets(writer, training_selection_data)
        return

    # LEGACY FORMATS: Single Training_Case_Selection sheet
    # Extract data
    selected_case_ids = training_selection_data['selected_case_ids']
    case_details = training_selection_data['case_details']
    metadata_df = training_selection_data.get('metadata_df', None)
    analysis_mode = training_selection_data.get('mode', None)  # 'concat', 'position', or None

    # Detect format: concat mode has 'cluster_info', multi-structure has 'structure_details'
    is_concat_mode = 'cluster_info' in training_selection_data

    if is_concat_mode:
        # Concat mode format
        cluster_info = training_selection_data['cluster_info']
        n_clusters = len(cluster_info)
        cluster_ids = ', '.join([str(cid) for cid in sorted(cluster_info.keys())])

        # Build Selected Cases table (concat mode: simpler format)
        selected_cases_data = []
        for case_id in selected_case_ids:
            details = case_details[case_id]
            case_data = {'Case_ID': case_id}
            case_data.update(_extract_case_metadata(case_id, metadata_df))

            if metadata_df is not None and case_id in metadata_df.index:
                if analysis_mode == 'position' and 'Structure_Name' in metadata_df.columns:
                    case_data['Structure_Name'] = metadata_df.loc[case_id, 'Structure_Name']
                elif 'All_Structure_Names' in metadata_df.columns:
                    case_data['Structure_Names'] = metadata_df.loc[case_id, 'All_Structure_Names']

            case_data.update({
                'Cluster_ID': details['cluster_id'],
                'Cluster_Size': details['cluster_size'],
                'Distance_From_Centroid': round(details['distance'], 4)
            })
            selected_cases_data.append(case_data)

        selected_cases_df = pd.DataFrame(selected_cases_data)

        # Build Per-Cluster Details table
        cluster_data = []
        for cluster_id, info in sorted(cluster_info.items()):
            cluster_data.append({
                'Cluster_ID': cluster_id,
                'Total_Size': info['total_size'],
                'Cases_Selected': info['selected_count'],
                'Selected_Case_IDs': ', '.join([str(cid) for cid in info['selected_case_ids']])
            })

        details_df = pd.DataFrame(cluster_data)
        details_header = 'PER-CLUSTER DETAILS'

        # Summary section
        summary_rows = [
            ['TRAINING CASE SELECTION SUMMARY', ''],
            ['', ''],
            ['Total training cases selected:', len(selected_case_ids)],
            ['Clusters analyzed:', n_clusters],
            ['Cluster IDs:', cluster_ids],
            ['', ''],
        ]

    else:
        # Multi-structure mode format (legacy)
        structure_details = training_selection_data['structure_details']
        n_structures = len(structure_details)
        structure_names = ', '.join(sorted(structure_details.keys()))

        # Build Selected Cases table (multi-structure format)
        selected_cases_data = []
        for case_id in selected_case_ids:
            details = case_details[case_id]
            case_data = {'Case_ID': case_id}
            case_data.update(_extract_case_metadata(case_id, metadata_df))

            if metadata_df is not None and case_id in metadata_df.index:
                if 'All_Structure_Names' in metadata_df.columns:
                    case_data['Structure_Names'] = metadata_df.loc[case_id, 'All_Structure_Names']

            case_data.update({
                'Structures_Represented': ', '.join(details['structures']),
                'Structure_Count': details['count']
            })
            selected_cases_data.append(case_data)

        selected_cases_df = pd.DataFrame(selected_cases_data)

        # Build Per-Structure Details table
        structure_data = []
        for structure_name, metadata in sorted(structure_details.items()):
            structure_data.append({
                'Structure': structure_name,
                'Clusters': metadata['clusters'],
                'Silhouette_Score': f"{metadata['silhouette']:.3f}",
                'Cases_Selected': metadata['cases_selected'],
                'Selected_Case_IDs': ', '.join([str(cid) for cid in metadata['case_ids']])
            })

        details_df = pd.DataFrame(structure_data)
        details_header = 'PER-STRUCTURE DETAILS'

        # Summary section
        summary_rows = [
            ['TRAINING CASE SELECTION SUMMARY', ''],
            ['', ''],
            ['Total training cases selected:', len(selected_case_ids)],
            ['Structures analyzed:', n_structures],
            ['Structure names:', structure_names],
            ['', ''],
        ]

    # Write summary section
    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_excel(writer, sheet_name='Training_Case_Selection',
                        index=False, header=False, startrow=0)

    # Write "SELECTED CASES" header
    start_row_cases = len(summary_rows)
    pd.DataFrame([['SELECTED CASES']]).to_excel(
        writer, sheet_name='Training_Case_Selection',
        index=False, header=False, startrow=start_row_cases
    )

    # Write selected cases table
    start_row_table = start_row_cases + 2
    selected_cases_df.to_excel(writer, sheet_name='Training_Case_Selection',
                               index=False, startrow=start_row_table)

    # Write details header (Per-Cluster or Per-Structure)
    start_row_details_header = start_row_table + len(selected_cases_df) + 3
    pd.DataFrame([[details_header]]).to_excel(
        writer, sheet_name='Training_Case_Selection',
        index=False, header=False, startrow=start_row_details_header
    )

    # Write details table
    start_row_details = start_row_details_header + 2
    details_df.to_excel(writer, sheet_name='Training_Case_Selection',
                       index=False, startrow=start_row_details)

    # Section 4: Detailed Cluster Mapping (optional)
    cluster_mapping = training_selection_data.get('cluster_mapping', [])
    if cluster_mapping:
        # Create DataFrame from cluster mapping
        cluster_mapping_df = pd.DataFrame(cluster_mapping)

        # Sort appropriately based on format
        if is_concat_mode:
            # Concat mode: only has Cluster_ID
            cluster_mapping_df = cluster_mapping_df.sort_values(
                ['Cluster_ID', 'Case_ID']
            ).reset_index(drop=True)
        else:
            # Multi-structure mode: has Structure column
            cluster_mapping_df = cluster_mapping_df.sort_values(
                ['Structure', 'Cluster_ID', 'Case_ID']
            ).reset_index(drop=True)

        # Round distance for readability
        cluster_mapping_df['Distance_From_Centroid'] = cluster_mapping_df['Distance_From_Centroid'].round(4)

        # Write "DETAILED CLUSTER MAPPING" header
        start_row_mapping_header = start_row_details + len(details_df) + 3
        pd.DataFrame([['DETAILED CLUSTER MAPPING']]).to_excel(
            writer, sheet_name='Training_Case_Selection',
            index=False, header=False, startrow=start_row_mapping_header
        )

        # Write cluster mapping table
        start_row_mapping = start_row_mapping_header + 2
        cluster_mapping_df.to_excel(writer, sheet_name='Training_Case_Selection',
                                    index=False, startrow=start_row_mapping)

        print(f"    ✓ Added detailed cluster mapping: {len(cluster_mapping_df)} entries")

    entity_type = "clusters" if is_concat_mode else "structures"
    entity_count = n_clusters if is_concat_mode else n_structures
    print(f"    ✓ Training_Case_Selection: {len(selected_case_ids)} cases selected from {entity_count} {entity_type}")


def _apply_conditional_formatting(excel_path, n_clusters):
    """Apply color coding to clustered data sheets."""
    apply_cluster_conditional_formatting(
        excel_path,
        n_clusters=n_clusters,
        main_sheet_name='Clustered_Data',
        cluster_column_name='Cluster'
    )


# =============================================================================
# Excel Formatting Utilities (merged from excel_utils.py)
# =============================================================================

def generate_cluster_summary(
    results_df: pd.DataFrame,
    feature_names: List[str],
    cluster_column: str = 'Cluster',
    n_top_features: int = 10,
    include_confidence: bool = False
) -> pd.DataFrame:
    """
    Generate standardized cluster summary statistics.

    Creates a summary DataFrame with cluster sizes, percentages, and optionally
    feature means and confidence scores.

    Args:
        results_df: DataFrame containing cluster assignments and features
        feature_names: List of feature column names to summarize
        cluster_column: Name of cluster assignment column (default: 'Cluster')
        n_top_features: Number of top features to include in summary
        include_confidence: Whether to include average confidence scores

    Returns:
        Summary DataFrame with cluster statistics
    """
    summary_data = []
    n_clusters = results_df[cluster_column].max() + 1
    total_cases = len(results_df)

    for cluster in range(n_clusters):
        mask = results_df[cluster_column] == cluster
        cluster_features = results_df[mask][feature_names]
        n_cases = mask.sum()

        summary_row = {
            cluster_column: cluster,
            'N_Cases': int(n_cases),
            'Percentage': f"{n_cases / total_cases * 100:.1f}%"
        }

        # Add average confidence score if requested
        if include_confidence and 'Confidence_Score' in results_df.columns:
            avg_confidence = results_df[mask]['Confidence_Score'].mean()
            summary_row['Avg_Confidence'] = f"{avg_confidence:.3f}"

        # Add feature statistics for top features
        if len(cluster_features) > 0:
            _add_feature_statistics(
                summary_row,
                cluster_features,
                results_df,
                feature_names,
                n_top_features
            )

        summary_data.append(summary_row)

    return pd.DataFrame(summary_data)


def _add_feature_statistics(
    summary_row: Dict[str, Any],
    cluster_features: pd.DataFrame,
    full_df: pd.DataFrame,
    feature_names: List[str],
    n_top: int
):
    """
    Add feature mean statistics to summary row.

    Internal helper to calculate and add feature means for top N features.
    Features are ranked by variance across clusters.
    """
    # Calculate feature means for this cluster
    feature_means = cluster_features.mean()

    # Rank features by variance across all clusters
    feature_variances = full_df[feature_names].var()
    top_features = feature_variances.nlargest(n_top).index.tolist()

    # Add top feature means to summary
    for feature in top_features[:n_top]:
        if feature in feature_means.index:
            summary_row[f"{feature}_mean"] = f"{feature_means[feature]:.3f}"


def apply_cluster_conditional_formatting(
    excel_path: Path,
    n_clusters: int,
    main_sheet_name: str = 'Clustered_Data',
    cluster_column_name: str = 'Cluster'
):
    """
    Apply color coding to clustered data Excel sheets.

    Applies consistent conditional formatting to cluster analysis output:
    - Main sheet: Color-codes cluster column
    - Cluster sheets: Applies cluster-specific header colors
    - Summary sheet: Color-codes cluster names

    Args:
        excel_path: Path to Excel file to format
        n_clusters: Number of clusters (for color palette)
        main_sheet_name: Name of main data sheet
        cluster_column_name: Name of cluster column to format
    """
    wb = load_workbook(excel_path)

    try:
        _format_main_sheet(wb, main_sheet_name, cluster_column_name, n_clusters)
        _format_cluster_sheets(wb, n_clusters)
        _format_summary_sheet(wb, n_clusters)
        wb.save(excel_path)
    finally:
        wb.close()


def _format_main_sheet(
    wb,
    sheet_name: str,
    cluster_col: str,
    n_clusters: int
):
    """Format main data sheet with cluster colors."""
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]

    # Find cluster column
    cluster_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == cluster_col:
            cluster_col_idx = idx
            break

    if cluster_col_idx is None:
        return

    # Apply color to each cluster value
    col_letter = get_column_letter(cluster_col_idx)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws[f"{col_letter}{row_idx}"]
        if cell.value is not None:
            cluster_num = int(cell.value)
            if cluster_num < len(CLUSTER_COLORS):
                cell.fill = PatternFill(
                    start_color=CLUSTER_COLORS[cluster_num].replace('#', ''),
                    end_color=CLUSTER_COLORS[cluster_num].replace('#', ''),
                    fill_type='solid'
                )


def _format_cluster_sheets(wb, n_clusters: int):
    """Format individual cluster data sheets."""
    for cluster in range(n_clusters):
        sheet_name = f'Data_Cluster_{cluster}'
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        color = CLUSTER_COLORS[cluster].replace('#', '')

        # Apply cluster color to header row
        for cell in ws[1]:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )
            cell.font = Font(bold=True, color='FFFFFF')


def _format_summary_sheet(wb, n_clusters: int):
    """Format summary sheet with cluster colors."""
    if 'Summary' not in wb.sheetnames:
        return

    ws = wb['Summary']

    # Find Cluster column
    cluster_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value and 'Cluster' in str(cell.value):
            cluster_col_idx = idx
            break

    if cluster_col_idx is None:
        return

    # Apply colors to cluster column
    col_letter = get_column_letter(cluster_col_idx)
    for row_idx in range(2, min(ws.max_row + 1, n_clusters + 2)):
        cell = ws[f"{col_letter}{row_idx}"]
        if cell.value is not None:
            cluster_num = int(cell.value)
            if cluster_num < len(CLUSTER_COLORS):
                cell.fill = PatternFill(
                    start_color=CLUSTER_COLORS[cluster_num].replace('#', ''),
                    end_color=CLUSTER_COLORS[cluster_num].replace('#', ''),
                    fill_type='solid'
                )
