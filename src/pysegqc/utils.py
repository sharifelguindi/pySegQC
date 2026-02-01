"""
Utility functions and constants for radiomics clustering analysis.

This module contains helper functions for feature filtering, volume normalization,
structure position detection, and view scan URL extraction.
"""

import re
import logging
from typing import List, Tuple, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Color palette for cluster visualization (up to 10 clusters)
CLUSTER_COLORS = [
    'C5E1A5',  # Light green
    'FFE082',  # Light yellow
    'FFAB91',  # Light orange
    '90CAF9',  # Light blue
    'CE93D8',  # Light purple
    'F48FB1',  # Light pink
    'B0BEC5',  # Light grey
    '80DEEA',  # Light cyan
    'FFCC80',  # Light amber
    'A5D6A7',  # Light mint
]

# ============================================================================
# FEATURE CLASSIFICATION: Volume-Dependent vs Volume-Independent
# ============================================================================

# CRITICAL: Volume-dependent features scale with structure size and will
# introduce size-based clustering even after StandardScaler normalization.
# These features remain correlated through power-law relationships:
#   - Linear dimensions scale as volume^(1/3)
#   - Surface areas scale as volume^(2/3)
#   - Volumes scale as volume^1
# Solution: Either EXCLUDE these features or NORMALIZE by volume before PCA.

VOLUME_DEPENDENT_FEATURES = [
    # Direct volume/size measurements
    'volume_cc',
    'voxel_count',
    'surface_area_mm2',

    # Bounding box dimensions (linear, scale with volume^1/3)
    'bbox_width',
    'bbox_height',
    'bbox_depth',
    'bbox_min_x',
    'bbox_min_y',
    'bbox_min_z',
    'bbox_max_x',
    'bbox_max_y',
    'bbox_max_z',

    # PCA axis lengths (linear dimensions)
    'pca_major_axis_length',
    'pca_minor_axis_length',
    'pca_least_axis_length',

    # Maximum diameter (linear dimension)
    'maximum_3d_diameter',

    # Convex hull metrics (volume and surface area)
    'convex_hull_volume',
    'convex_hull_surface_area',

    # Energy (sum over all voxels, scales with volume)
    'hu_energy',
]

VOLUME_INDEPENDENT_FEATURES = [
    # Shape ratios (dimensionless, volume cancels out)
    'bbox_aspect_ratio_xy',
    'bbox_aspect_ratio_xz',
    'bbox_aspect_ratio_yz',
    'pca_flatness',
    'elongation',

    # Normalized shape descriptors
    'sphericity',
    'compactness',
    'solidity',
    'surface_irregularity',
    'surface_to_volume_ratio',  # Note: inversely related to volume, but normalized

    # Intensity statistics (averaged over voxels, volume-independent)
    'hu_mean',
    'hu_std',
    'hu_min',
    'hu_max',
    'hu_median',
    'hu_percentile_10',
    'hu_percentile_30',
    'hu_percentile_50',
    'hu_percentile_70',
    'hu_percentile_90',
    'hu_skewness',
    'hu_kurtosis',
    'hu_entropy',
    'hu_interquartile_range',
    'hu_robust_mean_abs_deviation',

    # Texture features (normalized co-occurrence matrices)
    'glcm_contrast',
    'glcm_correlation',
    'glcm_energy',
    'glcm_homogeneity',
]

SPATIAL_POSITION_FEATURES = [
    # Anatomical location (not size-related, but may correlate with anatomy)
    'centroid_x',
    'centroid_y',
    'centroid_z',
]


def filter_volume_dependent_features(feature_names: List[str], exclude_spatial: bool = False) -> List[str]:
    """
    Filter out volume-dependent features from feature list.

    This ensures clustering is based on SHAPE/TEXTURE rather than SIZE.

    Args:
        feature_names: List of feature column names (e.g., ['001_volume_cc', '001_sphericity', ...])
        exclude_spatial: If True, also exclude spatial position features (centroid_x/y/z)

    Returns:
        List of volume-independent feature names

    Example:
        >>> features = ['001_volume_cc', '001_sphericity', '001_bbox_width']
        >>> filter_volume_dependent_features(features)
        ['001_sphericity']  # Only shape descriptor retained
    """
    logger.info("Filtering volume-dependent features...")

    volume_independent = []
    excluded_features = []

    for feat in feature_names:
        # Extract base feature name (remove position prefix like '001_' if present)
        # Position prefix pattern: 3 digits followed by underscore (e.g., '001_', '002_')
        if re.match(r'^\d{3}_', feat):
            base_feat = feat[4:]  # Remove first 4 characters '001_'
        else:
            base_feat = feat  # No position prefix, use as-is

        # Check if volume-dependent
        if base_feat in VOLUME_DEPENDENT_FEATURES:
            excluded_features.append(base_feat)
            continue

        # Check if spatial (optional exclusion)
        if exclude_spatial and base_feat in SPATIAL_POSITION_FEATURES:
            excluded_features.append(base_feat)
            continue

        # Keep volume-independent features
        volume_independent.append(feat)

    logger.info(f"✓ Filtered features: kept {len(volume_independent)}/{len(feature_names)}")
    logger.info(f"  - Excluded {len(excluded_features)} volume-dependent features")
    if len(excluded_features) > 0:
        logger.info(f"  - Excluded: {', '.join(sorted(set(excluded_features)))}")
    logger.info(f"  - Volume-independent features include: shape ratios, intensity stats, texture")

    return volume_independent


def normalize_by_volume(features_df: pd.DataFrame, feature_names: List[str]) -> Tuple[pd.DataFrame, List[str]]:
    """
    Normalize volume-dependent features by appropriate powers of volume.

    Converts absolute measurements to relative shape descriptors:
    - Linear dimensions (width, height, axes) → normalized by volume^(1/3)
    - Surface areas → normalized by volume^(2/3)
    - Volume itself → excluded (redundant after normalization)

    Args:
        features_df: DataFrame with radiomics features
        feature_names: List of feature column names

    Returns:
        Tuple of (normalized_df, updated_feature_names)

    Mathematical basis:
        For a structure with volume V:
        - Linear dimension L scales as L ∝ V^(1/3)
        - Surface area A scales as A ∝ V^(2/3)
        Normalizing by these powers yields dimensionless shape descriptors.
    """
    logger.info("Normalizing volume-dependent features by volume...")

    normalized_df = features_df.copy()
    new_feature_names = []

    # Features to normalize by volume^(1/3) - LINEAR DIMENSIONS
    linear_features = [
        'bbox_width', 'bbox_height', 'bbox_depth',
        'pca_major_axis_length', 'pca_minor_axis_length', 'pca_least_axis_length',
        'maximum_3d_diameter'
    ]

    # Features to normalize by volume^(2/3) - SURFACE AREAS
    surface_features = [
        'surface_area_mm2',
        'convex_hull_surface_area'
    ]

    # Features to EXCLUDE entirely (volume itself, voxel_count, energy, bounding box positions)
    exclude_features = [
        'volume_cc', 'voxel_count', 'hu_energy',
        'bbox_min_x', 'bbox_min_y', 'bbox_min_z',
        'bbox_max_x', 'bbox_max_y', 'bbox_max_z',
        'convex_hull_volume'  # Redundant with volume_cc
    ]

    normalized_count = 0
    excluded_count = 0

    for feat in feature_names:
        # Extract position and base feature name
        if '_' in feat:
            parts = feat.split('_')
            position = parts[0]
            base_feat = '_'.join(parts[1:])
        else:
            position = ''
            base_feat = feat

        # Get volume column for this position
        volume_col = f"{position}_volume_cc" if position else "volume_cc"

        if volume_col not in features_df.columns:
            # No volume available, keep feature as-is
            new_feature_names.append(feat)
            normalized_df[feat] = features_df[feat]
            continue

        # Normalize linear dimensions by V^(1/3)
        if base_feat in linear_features:
            new_col = f"{feat}_norm" if position else f"{base_feat}_norm"
            normalized_df[new_col] = features_df[feat] / (features_df[volume_col] ** (1/3))
            new_feature_names.append(new_col)
            normalized_count += 1

        # Normalize surface areas by V^(2/3)
        elif base_feat in surface_features:
            new_col = f"{feat}_norm" if position else f"{base_feat}_norm"
            normalized_df[new_col] = features_df[feat] / (features_df[volume_col] ** (2/3))
            new_feature_names.append(new_col)
            normalized_count += 1

        # Exclude volume and related features
        elif base_feat in exclude_features:
            excluded_count += 1
            # Don't add to new_feature_names

        # Keep volume-independent features as-is
        else:
            new_feature_names.append(feat)
            normalized_df[feat] = features_df[feat]

    # Keep only the new feature columns
    normalized_df = normalized_df[new_feature_names]

    logger.info(f"✓ Volume normalization complete:")
    logger.info(f"  - Normalized {normalized_count} features (linear dims, surface areas)")
    logger.info(f"  - Excluded {excluded_count} redundant features (volume, voxel_count, energy)")
    logger.info(f"  - Kept {len(new_feature_names) - normalized_count} volume-independent features as-is")

    return normalized_df, new_feature_names


def detect_structure_positions(feature_cols):
    """
    Detect structure positions from column names (e.g., 001_, 002_, ...).

    Args:
        feature_cols: List of feature column names

    Returns:
        tuple: (sorted_positions, n_features_per_structure)
    """
    positions = set()
    pattern = re.compile(r'^(\d{3})_')

    for col in feature_cols:
        match = pattern.match(col)
        if match:
            positions.add(int(match.group(1)))

    if not positions:
        raise ValueError("No structure position columns found (expected format: 001_feature_name)")

    sorted_positions = sorted(positions)
    n_structures = len(sorted_positions)

    # Count features for first position to determine features per structure
    first_pos = f"{sorted_positions[0]:03d}_"
    n_features_per_structure = sum(1 for col in feature_cols if col.startswith(first_pos))

    return sorted_positions, n_features_per_structure


def extract_view_scan_urls(excel_path):
    """
    Extract hyperlink URLs from the 'view_scan' column in the 'results' sheet.

    Args:
        excel_path: Path to Excel file

    Returns:
        dict: Mapping from row index to View Scan URL, or None if column doesn't exist
    """
    try:
        wb = load_workbook(excel_path)

        # Check if results sheet exists
        if 'results' not in wb.sheetnames:
            return None

        ws = wb['results']

        # Find view_scan column index
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        view_scan_col_idx = None
        index_col_idx = None

        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'view_scan':
                view_scan_col_idx = idx
            elif cell.value == 'index':
                index_col_idx = idx

        if not view_scan_col_idx:
            return None

        # Extract URLs from hyperlinks
        urls = {}
        for row_idx in range(2, ws.max_row + 1):
            # Get the index value (row number from PCA_Data)
            if index_col_idx:
                index_cell = ws.cell(row=row_idx, column=index_col_idx)
                index_value = index_cell.value
            else:
                index_value = row_idx - 1  # Fallback: use row number

            # Get the view_scan cell
            view_scan_cell = ws.cell(row=row_idx, column=view_scan_col_idx)

            # Extract hyperlink if it exists
            if view_scan_cell.hyperlink:
                urls[index_value] = view_scan_cell.hyperlink.target
            elif view_scan_cell.value and isinstance(view_scan_cell.value, str) and view_scan_cell.value.startswith('http'):
                # Fallback: if cell contains a URL string
                urls[index_value] = view_scan_cell.value

        wb.close()
        return urls if urls else None

    except Exception as e:
        print(f"⚠️  Warning: Could not extract view_scan URLs: {e}")
        return None


def get_plotly_click_handler_script():
    """
    Returns JavaScript code for handling plotly point clicks to navigate to View Scan URLs.

    This version works for standalone HTML files generated by Plotly's write_html().
    It dynamically finds all Plotly graph divs and attaches click handlers.

    Returns:
        str: JavaScript code as a string
    """
    return """
    <script>
    // Handle clicks on all Plotly plots to open View Scan URLs
    document.addEventListener('DOMContentLoaded', function() {
        var plots = document.querySelectorAll('.plotly-graph-div');
        plots.forEach(function(plot) {
            plot.on('plotly_click', function(data){
                var point = data.points[0];
                var url = point.customdata;
                if (url && url !== '' && url !== 'undefined' && url !== 'None') {
                    window.open(url, '_blank');
                }
            });
        });
    });
    </script>
    """


def get_dashboard_click_handler_script():
    """
    Returns JavaScript code for handling plotly point clicks in the dashboard.

    This version targets the specific div IDs used in the dashboard template
    (pca-2d-plot, pca-3d-plot) and works with embedded Plotly plots.

    Returns:
        str: JavaScript code as a string
    """
    return """
    <script>
    // Handle clicks on dashboard embedded plots to open View Scan URLs
    document.addEventListener('DOMContentLoaded', function() {
        ['pca-2d-plot', 'pca-3d-plot'].forEach(function(id) {
            var plotDiv = document.getElementById(id);
            if (plotDiv) {
                plotDiv.on('plotly_click', function(data){
                    var point = data.points[0];
                    var url = point.customdata;
                    if (url && url !== '' && url !== 'undefined' && url !== 'None') {
                        window.open(url, '_blank');
                    }
                });
            }
        });
    });
    </script>
    """


# =============================================================================
# Metadata Utilities (merged from metadata_utils.py)
# =============================================================================

def get_case_id(metadata_df: pd.DataFrame, index: int) -> str:
    """
    Extract best available case identifier from metadata.

    Checks Case_ID, MRN, and Patient_ID columns in order, falling back
    to the DataFrame index if none are available.

    Args:
        metadata_df: DataFrame with case metadata
        index: Row index to look up

    Returns:
        String case identifier
    """
    for col in ('Case_ID', 'MRN', 'Patient_ID'):
        if col in metadata_df.columns:
            val = metadata_df.iloc[index].get(col)
            if val is not None and pd.notna(val):
                return str(val)
    return str(metadata_df.index[index])


def extract_case_metadata(
    case_id: int,
    metadata_df: Optional[pd.DataFrame],
    metadata_columns: List[str] = [
        'MRN', 'Plan_ID', 'Session_ID', 'View_Scan_URL',
        'All_Structure_Names', 'Structure_Name'
    ]
) -> Dict[str, Any]:
    """
    Extract standardized metadata for a case.

    Builds a metadata dictionary for a specific case, extracting only
    the specified columns that exist in the DataFrame. Always includes
    Case_ID even if metadata_df is None.

    Args:
        case_id: Integer case identifier (DataFrame index)
        metadata_df: Optional DataFrame containing case metadata
        metadata_columns: List of column names to extract (default: standard columns)

    Returns:
        Dictionary with Case_ID and available metadata fields
    """
    case_metadata = {'Case_ID': case_id}

    if metadata_df is None or case_id not in metadata_df.index:
        return case_metadata

    for col in metadata_columns:
        if col in metadata_df.columns:
            case_metadata[col] = metadata_df.loc[case_id, col]

    return case_metadata


def print_banner(text: str, width: int = 70):
    """
    Print centered banner with borders.

    Displays text between two lines of equals signs. If text is shorter
    than width, it will be centered. If longer, it will not be truncated.

    Args:
        text: Banner text to display
        width: Banner width in characters (default: 70)
    """
    print(f"\n{'='*width}")
    if len(text) < width:
        print(text.center(width))
    else:
        print(text)
    print(f"{'='*width}")
