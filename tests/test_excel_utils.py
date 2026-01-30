import pytest
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from pysegqc.export import (
    apply_cluster_conditional_formatting,
    generate_cluster_summary
)


@pytest.fixture
def sample_clustered_excel(tmp_path):
    """Create sample Excel file with clustered data."""
    excel_path = tmp_path / "clustered.xlsx"

    # Create main sheet
    df = pd.DataFrame({
        'Case_ID': range(20),
        'Cluster': [0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 0, 1, 2, 0, 1],
        'Feature_1': np.random.randn(20),
        'Feature_2': np.random.randn(20)
    })

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Clustered_Data', index=False)

        # Create cluster sheets
        for cluster in range(3):
            cluster_df = df[df['Cluster'] == cluster]
            cluster_df.to_excel(writer, sheet_name=f'Data_Cluster_{cluster}', index=False)

        # Create summary
        summary = pd.DataFrame({
            'Cluster': [0, 1, 2],
            'N_Cases': [7, 7, 6],
            'Percentage': ['35.0%', '35.0%', '30.0%']
        })
        summary.to_excel(writer, sheet_name='Summary', index=False)

    return excel_path


def test_generate_cluster_summary_basic():
    """Test basic cluster summary generation."""
    df = pd.DataFrame({
        'Case_ID': range(10),
        'Cluster': [0, 0, 0, 1, 1, 1, 1, 2, 2, 2],
        'Feature_A': [1.0, 1.1, 0.9, 5.0, 5.2, 4.8, 5.1, 10.0, 9.8, 10.2],
        'Feature_B': [2.0, 2.1, 1.9, 6.0, 6.1, 5.9, 6.2, 8.0, 7.9, 8.1]
    })

    summary = generate_cluster_summary(
        df,
        feature_names=['Feature_A', 'Feature_B'],
        n_top_features=2
    )

    assert len(summary) == 3
    assert summary['Cluster'].tolist() == [0, 1, 2]
    assert summary['N_Cases'].tolist() == [3, 4, 3]
    assert '30.0%' in summary['Percentage'].iloc[0]


def test_generate_cluster_summary_with_confidence():
    """Test summary with confidence scores."""
    df = pd.DataFrame({
        'Cluster': [0, 0, 1, 1],
        'Confidence_Score': [0.95, 0.87, 0.92, 0.88],
        'Feature_1': [1, 2, 3, 4]
    })

    summary = generate_cluster_summary(
        df,
        feature_names=['Feature_1'],
        include_confidence=True
    )

    assert 'Avg_Confidence' in summary.columns
    assert '0.910' in summary.loc[0, 'Avg_Confidence']  # (0.95 + 0.87) / 2


def test_generate_cluster_summary_custom_column():
    """Test summary with custom cluster column name."""
    df = pd.DataFrame({
        'Predicted_Cluster': [0, 0, 1],
        'Feature_1': [1, 2, 3]
    })

    summary = generate_cluster_summary(
        df,
        feature_names=['Feature_1'],
        cluster_column='Predicted_Cluster'
    )

    assert 'Predicted_Cluster' in summary.columns


def test_apply_cluster_conditional_formatting(sample_clustered_excel):
    """Test Excel conditional formatting application."""
    apply_cluster_conditional_formatting(
        sample_clustered_excel,
        n_clusters=3,
        main_sheet_name='Clustered_Data',
        cluster_column_name='Cluster'
    )

    wb = load_workbook(sample_clustered_excel)

    # Verify sheets exist
    assert 'Clustered_Data' in wb.sheetnames
    assert 'Data_Cluster_0' in wb.sheetnames

    # Verify formatting applied (check fill colors exist)
    main_sheet = wb['Clustered_Data']
    # Row 2 should have fill (header is row 1)
    assert main_sheet['B2'].fill is not None

    wb.close()


def test_apply_formatting_missing_sheet(tmp_path):
    """Test graceful handling of missing sheet."""
    excel_path = tmp_path / "empty.xlsx"
    df = pd.DataFrame({'A': [1]})
    df.to_excel(excel_path, index=False)

    # Should not raise error
    apply_cluster_conditional_formatting(
        excel_path,
        n_clusters=3,
        main_sheet_name='NonExistent'
    )
