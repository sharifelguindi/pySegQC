"""
Shared test fixtures for pySegQC test suite.

Provides synthetic radiomics data generators and common test utilities.
"""

import pytest
import pandas as pd
import numpy as np
import joblib
from pathlib import Path


@pytest.fixture
def synthetic_radiomics_excel(tmp_path):
    """
    Generate synthetic radiomics Excel file with multi-structure data.

    Creates realistic radiomics features with 3 distinct cluster patterns:
    - Cluster 0: Spherical structures (high sphericity, low elongation)
    - Cluster 1: Elongated structures (low sphericity, high elongation)
    - Cluster 2: Irregular structures (medium values)

    Returns:
        Path: Path to temporary Excel file with 'PCA_Data' sheet
    """
    np.random.seed(42)
    n_samples = 50
    n_structures = 2

    # Metadata
    metadata = {
        'MRN': [f'MRN{i:04d}' for i in range(n_samples)],
        'Plan_ID': [f'PLAN{i:04d}' for i in range(n_samples)],
        'Session_ID': [f'SES{i:04d}' for i in range(n_samples)],
    }

    # Generate features for each structure
    features = {}
    for struct_num in range(1, n_structures + 1):
        prefix = f'{struct_num:03d}_'

        # Volume-independent shape features (3 clusters with distinct shapes)
        cluster_idx = np.random.choice([0, 1, 2], size=n_samples, p=[0.4, 0.35, 0.25])

        # Cluster 0: Spherical (high sphericity, low elongation)
        # Cluster 1: Elongated (low sphericity, high elongation)
        # Cluster 2: Irregular (medium sphericity, medium elongation)
        sphericity_means = [0.85, 0.45, 0.65]
        elongation_means = [1.2, 2.8, 1.8]

        sphericity = np.array([np.random.normal(sphericity_means[c], 0.05) for c in cluster_idx])
        elongation = np.array([np.random.normal(elongation_means[c], 0.2) for c in cluster_idx])

        features[f'{prefix}original_shape_Sphericity'] = np.clip(sphericity, 0, 1)
        features[f'{prefix}original_shape_Elongation'] = np.clip(elongation, 1, 5)
        features[f'{prefix}original_shape_Flatness'] = np.random.uniform(1.0, 3.0, n_samples)

        # Volume-dependent features (should be filtered in volume-independent mode)
        volumes = np.random.lognormal(3.5, 0.8, n_samples)
        features[f'{prefix}original_shape_VoxelVolume'] = volumes
        features[f'{prefix}original_shape_SurfaceArea'] = volumes ** (2/3) * np.random.uniform(4.5, 5.5, n_samples)
        features[f'{prefix}original_shape_Maximum3DDiameter'] = volumes ** (1/3) * np.random.uniform(1.5, 2.5, n_samples)

        # Intensity features (volume-independent)
        features[f'{prefix}original_firstorder_Mean'] = np.random.normal(50, 20, n_samples)
        features[f'{prefix}original_firstorder_Median'] = np.random.normal(48, 18, n_samples)
        features[f'{prefix}original_firstorder_StandardDeviation'] = np.random.uniform(10, 30, n_samples)

        # Texture features
        features[f'{prefix}original_glcm_Contrast'] = np.random.uniform(20, 100, n_samples)
        features[f'{prefix}original_glcm_Correlation'] = np.random.uniform(0.4, 0.9, n_samples)
        features[f'{prefix}original_glcm_Energy'] = np.random.uniform(0.01, 0.1, n_samples)

    # Combine metadata and features
    df = pd.DataFrame({**metadata, **features})

    # Save to Excel
    excel_path = tmp_path / 'synthetic_radiomics.xlsx'
    df.to_excel(excel_path, sheet_name='PCA_Data', index=False)

    return excel_path


@pytest.fixture
def temp_output_dir(tmp_path):
    """Create temporary output directory for tests"""
    output_dir = tmp_path / 'output'
    output_dir.mkdir()
    return output_dir


@pytest.fixture
def trained_models_dir(tmp_path, synthetic_radiomics_excel):
    """
    Create trained models directory with serialized models.

    Trains a simple model on synthetic data and saves it for prediction tests.

    Returns:
        Path: Directory containing trained_models.pkl
    """
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler
    from sklearn.cluster import AgglomerativeClustering

    # Load synthetic data using the actual data_loader to match behavior
    from pysegqc.data_loader import load_and_preprocess_data
    metadata, features_df, feature_names, structure_info = load_and_preprocess_data(
        synthetic_radiomics_excel, 'PCA_Data', mode='position', position=1
    )

    # Handle missing values with median imputation (preserve DataFrame for feature names)
    from sklearn.impute import SimpleImputer
    imputer = SimpleImputer(strategy='median')
    X_imputed_array = imputer.fit_transform(features_df)

    # Convert imputed data back to DataFrame to preserve feature names
    X_imputed_df = pd.DataFrame(
        X_imputed_array,
        columns=features_df.columns,
        index=features_df.index
    )

    # Create models (fit with DataFrame to track feature names correctly)
    scaler = StandardScaler()
    X_scaled_array = scaler.fit_transform(X_imputed_df)
    X_scaled_df = pd.DataFrame(
        X_scaled_array,
        columns=X_imputed_df.columns,
        index=X_imputed_df.index
    )

    pca_model = PCA(n_components=5)
    X_pca = pca_model.fit_transform(X_scaled_df)

    clustering_model = AgglomerativeClustering(n_clusters=3)
    clustering_model.fit(X_pca)

    # Calculate centroids from clustering
    centroids = np.array([X_pca[clustering_model.labels_ == i].mean(axis=0)
                          for i in range(clustering_model.n_clusters)])

    # Save models (match keys expected by prediction.py)
    # feature_names will have prefix stripped in position mode (e.g., 'Sphericity' not '001_Sphericity')
    models_dict = {
        'pca': pca_model,  # prediction.py expects 'pca' not 'pca_model'
        'pca_model': pca_model,  # Keep for backward compatibility
        'scaler': scaler,
        'centroids': centroids,
        'clustering_model': clustering_model,
        'feature_names': feature_names,  # Now matches data_loader output
        'n_clusters': clustering_model.n_clusters,
        'impute_strategy': 'median',
        'n_components': 5,
        'mode': 'position',
        'position': 1
    }

    models_path = tmp_path / 'trained_models.pkl'
    joblib.dump(models_dict, models_path)

    return tmp_path


@pytest.fixture
def sample_metadata_df():
    """Create sample metadata DataFrame for testing"""
    return pd.DataFrame({
        'MRN': [f'MRN{i:04d}' for i in range(30)],
        'Plan_ID': [f'PLAN{i:04d}' for i in range(30)],
        'Session_ID': [f'SES{i:04d}' for i in range(30)]
    })


@pytest.fixture
def sample_features_multistructure():
    """Create sample features DataFrame with position prefixes"""
    n_samples = 30
    return pd.DataFrame({
        '001_original_shape_Sphericity': np.random.uniform(0.4, 0.9, n_samples),
        '001_original_shape_Elongation': np.random.uniform(1.0, 3.0, n_samples),
        '002_original_shape_Sphericity': np.random.uniform(0.4, 0.9, n_samples),
        '002_original_shape_Elongation': np.random.uniform(1.0, 3.0, n_samples)
    })


@pytest.fixture
def three_cluster_data():
    """Generate data with 3 clear clusters for testing"""
    from sklearn.datasets import make_blobs
    X, labels = make_blobs(n_samples=150, n_features=5, centers=3,
                          cluster_std=0.5, random_state=42)
    return X, labels


@pytest.fixture
def synthetic_radiomics_with_results_sheet(tmp_path):
    """
    Generate synthetic radiomics Excel with results and Structure_Mapping sheets.

    This fixture creates a realistic Excel file with:
    - PCA_Data sheet (main radiomics features)
    - results sheet (database IDs and View_Scan_URL hyperlinks)
    - Structure_Mapping sheet (position to structure name mapping)

    Returns:
        Path: Path to temporary Excel file
    """
    np.random.seed(42)
    n_samples = 30
    n_structures = 2

    # Main radiomics data (simplified version)
    metadata = {
        'MRN': [f'MRN{i:04d}' for i in range(n_samples)],
        'Plan_ID': [f'PLAN{i:04d}' for i in range(n_samples)],
        'Session_ID': [f'SES{i:04d}' for i in range(n_samples)],
    }

    features = {}
    for struct_num in range(1, n_structures + 1):
        prefix = f'{struct_num:03d}_'
        features[f'{prefix}original_shape_Sphericity'] = np.random.uniform(0.4, 0.9, n_samples)
        features[f'{prefix}original_shape_Elongation'] = np.random.uniform(1.0, 3.0, n_samples)
        features[f'{prefix}original_shape_VoxelVolume'] = np.random.lognormal(3.5, 0.8, n_samples)

    pca_data_df = pd.DataFrame({**metadata, **features})

    # Results sheet with database IDs
    results_data = {
        'Plan_ID': [f'PLAN{i:04d}' for i in range(n_samples)],
        'Session_ID': [f'SES{i:04d}' for i in range(n_samples)],
        'DB_Plan_ID': [1000 + i for i in range(n_samples)],
        'DB_Session_ID': [2000 + i for i in range(n_samples)],
        'view_scan': [f'http://example.com/scan/{i}' for i in range(n_samples)],  # Will add hyperlinks
    }
    results_df = pd.DataFrame(results_data)

    # Structure mapping sheet
    mapping_data = {
        'position': ['001', '002'],
        'structure_name': ['CTV', 'Bladder']
    }
    mapping_df = pd.DataFrame(mapping_data)

    # Save to Excel with multiple sheets
    excel_path = tmp_path / 'synthetic_radiomics_with_results.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        pca_data_df.to_excel(writer, sheet_name='PCA_Data', index=False)
        results_df.to_excel(writer, sheet_name='results', index=False)
        mapping_df.to_excel(writer, sheet_name='Structure_Mapping', index=False)

    # Add hyperlinks to view_scan column using openpyxl
    from openpyxl import load_workbook
    from openpyxl.worksheet.hyperlink import Hyperlink
    wb = load_workbook(excel_path)
    ws = wb['results']

    # Find view_scan column
    view_scan_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'view_scan':
            view_scan_col = idx
            break

    if view_scan_col:
        for row in range(2, len(results_data['view_scan']) + 2):
            cell = ws.cell(row, view_scan_col)
            cell.hyperlink = cell.value

    wb.save(excel_path)
    wb.close()

    return excel_path


@pytest.fixture
def high_missingness_data(tmp_path):
    """
    Generate synthetic radiomics data with high missingness (>80% in some columns).

    This tests the data_loader's column removal logic for features with >80% missing values.

    Returns:
        Path: Path to temporary Excel file
    """
    np.random.seed(42)
    n_samples = 50

    metadata = {
        'MRN': [f'MRN{i:04d}' for i in range(n_samples)],
        'Plan_ID': [f'PLAN{i:04d}' for i in range(n_samples)],
    }

    # Create features with varying missingness levels
    features = {
        '001_good_feature_1': np.random.uniform(0.4, 0.9, n_samples),  # 0% missing
        '001_good_feature_2': np.random.uniform(1.0, 3.0, n_samples),  # 0% missing
        '001_sparse_feature': np.full(n_samples, np.nan),  # 100% missing - should be dropped
        '001_mostly_missing': np.full(n_samples, np.nan),  # 90% missing - should be dropped
    }

    # Add some non-NaN values to mostly_missing (10% present)
    features['001_mostly_missing'][:5] = np.random.uniform(0, 1, 5)

    df = pd.DataFrame({**metadata, **features})

    excel_path = tmp_path / 'high_missingness_data.xlsx'
    df.to_excel(excel_path, sheet_name='PCA_Data', index=False)

    return excel_path


@pytest.fixture
def large_radiomics_dataset(tmp_path):
    """
    Generate large synthetic radiomics dataset (>30 samples) for dendrogram truncation testing.

    The visualization module truncates dendrograms when n_samples > 30 to keep plots readable.

    Returns:
        Path: Path to temporary Excel file with 60 samples
    """
    np.random.seed(42)
    n_samples = 60  # Exceeds dendrogram truncation threshold of 30
    n_structures = 1  # Single structure for simplicity

    metadata = {
        'MRN': [f'MRN{i:04d}' for i in range(n_samples)],
        'Plan_ID': [f'PLAN{i:04d}' for i in range(n_samples)],
        'Session_ID': [f'SES{i:04d}' for i in range(n_samples)],
    }

    # Generate features with clear cluster patterns
    cluster_idx = np.random.choice([0, 1, 2], size=n_samples, p=[0.4, 0.35, 0.25])
    sphericity_means = [0.85, 0.45, 0.65]
    elongation_means = [1.2, 2.8, 1.8]

    features = {}
    for struct_num in range(1, n_structures + 1):
        prefix = f'{struct_num:03d}_'
        sphericity = np.array([np.random.normal(sphericity_means[c], 0.05) for c in cluster_idx])
        elongation = np.array([np.random.normal(elongation_means[c], 0.2) for c in cluster_idx])

        features[f'{prefix}original_shape_Sphericity'] = np.clip(sphericity, 0, 1)
        features[f'{prefix}original_shape_Elongation'] = np.clip(elongation, 1, 5)
        features[f'{prefix}original_shape_VoxelVolume'] = np.random.lognormal(3.5, 0.8, n_samples)
        features[f'{prefix}original_firstorder_Mean'] = np.random.normal(50, 20, n_samples)
        features[f'{prefix}original_glcm_Contrast'] = np.random.uniform(20, 100, n_samples)

    df = pd.DataFrame({**metadata, **features})

    excel_path = tmp_path / 'large_radiomics_dataset.xlsx'
    df.to_excel(excel_path, sheet_name='PCA_Data', index=False)

    return excel_path


@pytest.fixture
def synthetic_nifti_data(tmp_path):
    """
    Generate synthetic NIfTI image and mask files for testing.

    Creates paired directories structure:
        tmp_path/
        ├── images/
        │   ├── patient001_CT.nii.gz
        │   ├── patient002_CT.nii.gz
        │   └── patient003_CT.nii.gz
        └── masks/
            ├── patient001_mask.nii.gz
            ├── patient002_mask.nii.gz
            └── patient003_mask.nii.gz

    Images simulate CT scans with HU-like values, masks contain spherical ROIs.

    Returns:
        Dict with keys:
            - 'data_dir': Root directory
            - 'pairs': List of (image_path, mask_path) tuples
            - 'n_cases': Number of cases created
    """
    import nibabel as nib

    np.random.seed(42)

    # Create directory structure
    data_dir = tmp_path / 'nifti_data'
    images_dir = data_dir / 'images'
    masks_dir = data_dir / 'masks'
    images_dir.mkdir(parents=True)
    masks_dir.mkdir(parents=True)

    pairs = []
    n_cases = 3

    for case_id in range(1, n_cases + 1):
        # Create synthetic CT image (64x64x32 for speed)
        image_array = np.random.normal(-200, 100, (32, 64, 64)).astype(np.float32)

        # Add a "tumor" region with higher intensity
        center = (16, 32, 32)
        z, y, x = np.ogrid[:32, :64, :64]
        distance = np.sqrt((z - center[0])**2 + (y - center[1])**2 + (x - center[2])**2)
        tumor_mask = distance <= 10
        image_array[tumor_mask] += 250  # Tumor has higher HU values

        # Create spherical mask (binary)
        mask_array = (distance <= 10).astype(np.uint8)

        # Save image
        image_path = images_dir / f'patient{case_id:03d}_CT.nii.gz'
        affine = np.eye(4)
        affine[0:3, 0:3] = np.diag([1.5, 1.5, 3.0])  # Spacing: 1.5x1.5x3.0 mm
        nib.save(nib.Nifti1Image(image_array, affine), image_path)

        # Save mask
        mask_path = masks_dir / f'patient{case_id:03d}_mask.nii.gz'
        nib.save(nib.Nifti1Image(mask_array, affine), mask_path)

        pairs.append((image_path, mask_path))

    return {
        'data_dir': data_dir,
        'pairs': pairs,
        'n_cases': n_cases
    }


@pytest.fixture
def synthetic_multiclass_nifti(tmp_path):
    """
    Generate synthetic NIfTI with multi-class mask for testing.

    Creates:
        - Single CT image
        - Multi-class mask with 3 nested structures:
          * Class 1: Outer sphere (radius 15)
          * Class 2: Middle sphere (radius 10)
          * Class 3: Inner sphere (radius 5)

    Returns:
        Dict with keys:
            - 'image_path': Path to image NIfTI
            - 'mask_path': Path to multi-class mask NIfTI
            - 'n_classes': Number of classes (3)
    """
    import nibabel as nib

    np.random.seed(42)

    # Create directory
    data_dir = tmp_path / 'multiclass_nifti'
    data_dir.mkdir(parents=True)

    # Create synthetic CT image
    image_array = np.random.normal(-200, 100, (32, 64, 64)).astype(np.float32)

    # Create multi-class mask (nested spheres)
    center = (16, 32, 32)
    z, y, x = np.ogrid[:32, :64, :64]
    distance = np.sqrt((z - center[0])**2 + (y - center[1])**2 + (x - center[2])**2)

    mask_array = np.zeros((32, 64, 64), dtype=np.uint8)
    mask_array[distance <= 15] = 1  # Outer structure
    mask_array[distance <= 10] = 2  # Middle structure
    mask_array[distance <= 5] = 3   # Inner structure

    # Add corresponding intensities to image
    image_array[mask_array == 1] += 100
    image_array[mask_array == 2] += 200
    image_array[mask_array == 3] += 300

    # Save files
    affine = np.eye(4)
    affine[0:3, 0:3] = np.diag([1.5, 1.5, 3.0])

    image_path = data_dir / 'patient001_CT.nii.gz'
    mask_path = data_dir / 'patient001_multiclass.nii.gz'

    nib.save(nib.Nifti1Image(image_array, affine), image_path)
    nib.save(nib.Nifti1Image(mask_array, affine), mask_path)

    return {
        'image_path': image_path,
        'mask_path': mask_path,
        'n_classes': 3
    }
