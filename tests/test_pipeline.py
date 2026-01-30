"""Integration tests for pipeline module"""

import pytest
from pathlib import Path
from argparse import Namespace
from pysegqc.pipeline import run_analysis_pipeline


def test_full_pipeline_position_mode(synthetic_radiomics_excel, temp_output_dir):
    """Test complete pipeline in position mode"""
    args = Namespace(
        mode='position',
        position=1,
        n_components=5,
        n_clusters=3,
        auto_k=False,
        method='hierarchical',
        impute='median',
        volume_independent=False,
        volume_normalize=False,
        select_training_cases=False,
        max_k=6,
        sheet='PCA_Data'
    )

    result = run_analysis_pipeline(args, synthetic_radiomics_excel, temp_output_dir)

    # Check key outputs exist
    assert (temp_output_dir / 'synthetic_radiomics_clustered.xlsx').exists()
    # PCA loadings and cluster stats are now sheets inside the Excel file
    import openpyxl
    wb = openpyxl.load_workbook(temp_output_dir / 'synthetic_radiomics_clustered.xlsx')
    assert 'PCA_Loadings' in wb.sheetnames
    assert 'Cluster_Statistics' in wb.sheetnames
    wb.close()
    assert (temp_output_dir / 'analysis_dashboard.html').exists()
    assert (temp_output_dir / 'analysis_report.json').exists()
    assert (temp_output_dir / 'trained_models.pkl').exists()


def test_pipeline_with_auto_k(synthetic_radiomics_excel, temp_output_dir):
    """Test pipeline with automatic k selection"""
    args = Namespace(
        mode='position',
        position=1,
        n_components=5,
        n_clusters=None,
        auto_k=True,
        max_k=6,
        method='hierarchical',
        impute='median',
        volume_independent=False,
        volume_normalize=False,
        select_training_cases=False,
        sheet='PCA_Data'
    )

    result = run_analysis_pipeline(args, synthetic_radiomics_excel, temp_output_dir)

    # Should successfully complete with auto-selected k
    assert (temp_output_dir / 'synthetic_radiomics_clustered.xlsx').exists()


def test_pipeline_volume_independent_mode(synthetic_radiomics_excel, temp_output_dir):
    """Test pipeline with volume-independent filtering"""
    args = Namespace(
        mode='position',
        position=1,
        n_components=5,
        n_clusters=3,
        auto_k=False,
        method='hierarchical',
        impute='median',
        volume_independent=True,  # Enable volume filtering
        volume_normalize=False,
        select_training_cases=False,
        max_k=6,
        sheet='PCA_Data'
    )

    result = run_analysis_pipeline(args, synthetic_radiomics_excel, temp_output_dir)

    # Should filter volume-dependent features
    assert (temp_output_dir / 'synthetic_radiomics_clustered.xlsx').exists()


def test_pipeline_concat_mode(synthetic_radiomics_excel, temp_output_dir):
    """Test pipeline in concat mode (multi-structure)"""
    args = Namespace(
        mode='concat',
        position=None,
        n_components=5,
        n_clusters=3,
        auto_k=False,
        method='hierarchical',
        impute='median',
        volume_independent=False,
        volume_normalize=False,
        select_training_cases=False,
        max_k=6,
        sheet='PCA_Data'
    )

    result = run_analysis_pipeline(args, synthetic_radiomics_excel, temp_output_dir)

    # Should successfully complete in concat mode
    assert (temp_output_dir / 'synthetic_radiomics_clustered.xlsx').exists()
